# Copyright (C) 2024 Beksultan Artykbaev - All Rights Reserved

import os
import logging
import time
from typing import List, Union, Tuple, Any

import xlsxwriter
import openpyxl
# ---------------------------------------
# NOTE: Do not remove these pyexcel imports as it requires direct imports to successfully compile using pyinstaller
import pyexcel as p
import pyexcel_xls 
import pyexcel_xlsx
import pyexcel_xlsxw
import pyexcel_io
import pyexcel_io.writers
# ---------------------------------------

# Setting up logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(filename)s:%(lineno)d]: %(message)s')
logger = logging.getLogger(__name__)

def in_use(filename):
    '''Checks if program is currently being used (in run).'''
    try:
        os.rename(filename, filename)
        return False
    except:    
        return True

def generate_excel(matrix: List[List[Union[str, int]]],
                   output_filename: str, frequency_analysis: List[tuple[Any, int]]=None) -> None:
    '''Generate xlsx file from co-occurrence matrix.'''
    try:
        workbook = xlsxwriter.Workbook(output_filename)
        worksheet = workbook.add_worksheet("Co-occurrence matrix")
        col = 0

        for row, data in enumerate(matrix):
            worksheet.write_row(row, col, data)

        if frequency_analysis != None:
            second_worksheet = workbook.add_worksheet("Frequency Analysis")
            col = 0

            for row, data in enumerate(frequency_analysis):
                second_worksheet.write_row(row, col, data)

        workbook.close()
    except xlsxwriter.exceptions.FileCreateError:
        logger.info(f"You are trying to save to an opened file: {repr(output_filename)}, please close that file.")
        while in_use(output_filename):
            time.sleep(0.1)
        logger.info(f"You have closed: {repr(output_filename)}, continuing saving to that file.")
        generate_excel(matrix, output_filename, frequency_analysis)
            
def create_xlsx_copy(filename:str) -> None:
    '''Convert file with other spreadsheet filetype format to .xlsx'''
    if filename.endswith(".xls"):
        p.save_book_as(file_name=filename,
                    dest_file_name=f"{filename}x")

    elif filename.endswith(".csv"):
        p.save_book_as(file_name=filename,
                    dest_file_name=filename.replace(".csv", ".xlsx"))
    else:
        logger.warning(f'''You are trying to open file with type that is not tested. ({filename.split(".")[1]})
The tested formats are: (csv, xlsx, xls)''')
        new_filename = filename.split(".")[0] + ".xlsx"
        p.save_book_as(filename=filename, dest_file_name=new_filename)

def load_xls_sheet_values(xls_filepath: str, ranges: str, sheet_name=None, delimeter:str=";") -> List[List[str]]:
    '''Reads given XLS(X) specific sheet and returns values of cells in given range.'''
    # Converting xls file to .xlsx because openpyxl doesn't support xls
    is_temp = False
    if xls_filepath.split(".")[1] != "xlsx":
        is_temp = True
        create_xlsx_copy(xls_filepath)
        xls_filepath = xls_filepath.split(".")[0] + ".xlsx"
    # Loading workbook
    workbook = openpyxl.load_workbook(xls_filepath, True)
    
    # If specific sheet selected use it, active sheet otherwise
    if sheet_name != None:
        workbook.active = workbook[sheet_name]
    curr = workbook.active

    texts = list()

    # Extracting cell values
    for range_ in ranges.split("|"):
        for cell in curr[range_]:
            try:
                if cell[0].value != None:
                    texts.append(list())
                    for word in cell[0].value.split(delimeter):
                        texts[-1].append(word.strip())

            except TypeError as e:
                pass

    workbook.close()
    if is_temp:
        os.remove(xls_filepath)

    return texts

def get_active_sheetname(xls_filepath: str) -> List[str]:
    # Converting xls file to .xlsx because openpyxl doesn't support xls
    is_temp = False
    if xls_filepath.split(".")[1] != "xlsx":
        is_temp = True
        create_xlsx_copy(xls_filepath)
        xls_filepath = xls_filepath.split(".")[0] + ".xlsx"

    # Loading workbook
    workbook = openpyxl.load_workbook(xls_filepath, True)
    sheetname = workbook.active.title
    workbook.close()
    if is_temp:
        os.remove(xls_filepath)
    return sheetname

def read_savedrecs(filename: str) -> List[List[str]]:
    '''Read tab-delimeted file (WOS savedrecs.txt)'''
    with open(filename, encoding="utf-8-sig") as f:
        text = f.read()
    lines = text.split("\n")
    matrix = list()

    for line in lines:
        elements = line.split("\t")
        matrix.append(elements)
    return matrix